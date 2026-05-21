from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from appointments.models import Appointment
from services.models import ServiceRecord
from staff.models import Doctor
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from patients.models import Patient
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@login_required
def dashboard_router(request):
    user = request.user
    if user.is_superuser or user.is_admin():
        return redirect('dashboard:admin_dashboard')
    elif user.is_doctor():
        return redirect('staff:doctor_dashboard')
    elif user.is_receptionist():
        return redirect('appointments:index')
    else:
        return redirect('accounts:profile')

@login_required
@permission_required('staff.view_receptionist', raise_exception=True)
def admin_dashboard(request):

    # Filter logic
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    exam_type = request.GET.get('exam_type', 'all')
    
    today = timezone.now().date()
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = today
        
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = start_date # If only one date provided, it's a single day

    # Fetch data
    appointments = Appointment.objects.filter(date__date__range=[start_date, end_date]).exclude(status='cancelled')
    services = ServiceRecord.objects.filter(created_at__date__range=[start_date, end_date])
    
    # Summary Statistics
    total_app_revenue = appointments.aggregate(Sum('cost'))['cost__sum'] or 0
    total_service_revenue = services.aggregate(Sum('service__price'))['service__price__sum'] or 0
    total_revenue = total_app_revenue + total_service_revenue
    
    # Detailed Doctor Analytics
    doctors = Doctor.objects.all()
    if exam_type != 'all':
        doctors = doctors.filter(examination_type=exam_type)
        
    doctor_data = []
    
    total_doctor_share = 0
    total_clinic_share = 0
    
    for doctor in doctors:
        # Appointments for this doctor
        doc_apps = appointments.filter(doctor=doctor)
        doc_app_rev = doc_apps.aggregate(Sum('cost'))['cost__sum'] or 0
        
        # Services for this doctor
        doc_services = services.filter(doctor=doctor)
        doc_service_rev = doc_services.aggregate(Sum('service__price'))['service__price__sum'] or 0
        doc_service_share = doc_services.aggregate(Sum('doctor_amount'))['doctor_amount__sum'] or 0
        
        # Calculate Appointment Share
        if doctor.examination_type == 'percentage':
            percentage = doctor.percentage_value or 0
            doc_app_share = doc_app_rev * (percentage / 100)
        elif doctor.examination_type == 'time_share':
            price = doctor.price_value or 0
            doc_app_share = doc_apps.count() * price
        else:
            doc_app_share = doc_app_rev
            
        doc_total_share = doc_app_share + doc_service_share
        doc_clinic_share = (doc_app_rev + doc_service_rev) - doc_total_share
        
        total_doctor_share += doc_total_share
        total_clinic_share += doc_clinic_share
        
        if doc_apps.count() > 0 or doc_services.count() > 0:
            doctor_data.append({
                'doctor': doctor,
                'app_count': doc_apps.count(),
                'app_revenue': doc_app_rev,
                'service_revenue': doc_service_rev,
                'doctor_share': doc_total_share,
                'clinic_share': doc_clinic_share,
                'type': doctor.get_examination_type_display()
            })
            
    # Detailed Service Analytics
    service_filtered = services
    if exam_type != 'all':
        service_filtered = services.filter(doctor__examination_type=exam_type)
        
    service_details = service_filtered.values(
        'service__name', 
        'doctor__user__first_name', 
        'doctor__user__last_name'
    ).annotate(
        count=Count('id'),
        total_price=Sum('service__price'),
        total_doctor_share=Sum('doctor_amount'),
        total_clinic_share=Sum('clinic_amount')
    ).order_by('-total_price')

    # Detailed Appointment (Consultation) Analytics
    app_details_filtered = appointments
    if exam_type != 'all':
        app_details_filtered = appointments.filter(doctor__examination_type=exam_type)
    
    appointment_details = app_details_filtered.values(
        'doctor__user__first_name', 
        'doctor__user__last_name',
        'doctor__examination_type'
    ).annotate(
        count=Count('id'),
        total_revenue=Sum('cost'),
    ).order_by('-total_revenue')

    context = {
        'page_title': 'لوحة تحكم المدير',
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_revenue': total_revenue,
        'total_app_revenue': total_app_revenue,
        'total_service_revenue': total_service_revenue,
        'total_doctor_share': total_doctor_share,
        'total_clinic_share': total_clinic_share,
        'app_count': appointments.count(),
        'service_count': services.count(),
        'doctor_data': doctor_data,
        'service_details': service_details,
        'appointment_details': appointment_details,
        'exam_type': exam_type,
    }
    
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
@permission_required('patients.view_patient', raise_exception=True)
def unified_search(request):

    query = request.GET.get('q', '')
    search_type = request.GET.get('type', 'patient')  # default to patient
    page_number = request.GET.get('page', 1)

    results = []
    
    if search_type == 'doctor':
        queryset = Doctor.objects.select_related('user').all()
        if query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__phone_number__icontains=query) |
                Q(specialization__icontains=query)
            )
        results = queryset.order_by('user__first_name')
    else:  # patient
        queryset = Patient.objects.select_related('user').all()
        if query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__phone_number__icontains=query)
            )
        results = queryset.order_by('-created_at')

    paginator = Paginator(results, 10)  # 10 results per page
    page_obj = paginator.get_page(page_number)

    context = {
        'page_title': 'البحث الموحد',
        'page_obj': page_obj,
        'query': query,
        'search_type': search_type,
    }
    return render(request, 'dashboard/unified_search.html', context)


@login_required
@permission_required('staff.view_receptionist', raise_exception=True)
def export_excel(request):
    # Filter logic (same as admin_dashboard)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    exam_type = request.GET.get('exam_type', 'all')
    
    today = timezone.now().date()
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = today
        
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = start_date

    # Fetch data
    appointments = Appointment.objects.filter(date__date__range=[start_date, end_date]).exclude(status='cancelled')
    services = ServiceRecord.objects.filter(created_at__date__range=[start_date, end_date])
    
    if exam_type != 'all':
        appointments = appointments.filter(doctor__examination_type=exam_type)
        services = services.filter(doctor__examination_type=exam_type)
    
    # Calculate statistics
    total_app_revenue = appointments.aggregate(Sum('cost'))['cost__sum'] or 0
    total_service_revenue = services.aggregate(Sum('service__price'))['service__price__sum'] or 0
    total_revenue = total_app_revenue + total_service_revenue
    
    doctors = Doctor.objects.all()
    if exam_type != 'all':
        doctors = doctors.filter(examination_type=exam_type)
        
    doctor_data = []
    total_doctor_share = 0
    total_clinic_share = 0
    
    for doctor in doctors:
        doc_apps = appointments.filter(doctor=doctor)
        doc_app_rev = doc_apps.aggregate(Sum('cost'))['cost__sum'] or 0
        
        doc_services = services.filter(doctor=doctor)
        doc_service_rev = doc_services.aggregate(Sum('service__price'))['service__price__sum'] or 0
        doc_service_share = doc_services.aggregate(Sum('doctor_amount'))['doctor_amount__sum'] or 0
        
        if doctor.examination_type == 'percentage':
            percentage = doctor.percentage_value or 0
            doc_app_share = doc_app_rev * (percentage / 100)
        elif doctor.examination_type == 'time_share':
            price = doctor.price_value or 0
            doc_app_share = doc_apps.count() * price
        else:
            doc_app_share = doc_app_rev
            
        doc_total_share = doc_app_share + doc_service_share
        doc_clinic_share = (doc_app_rev + doc_service_rev) - doc_total_share
        
        total_doctor_share += doc_total_share
        total_clinic_share += doc_clinic_share
        
        if doc_apps.count() > 0 or doc_services.count() > 0:
            doctor_data.append({
                'doctor': doctor.user.get_full_name(),
                'type': doctor.get_examination_type_display(),
                'app_count': doc_apps.count(),
                'app_revenue': doc_app_rev,
                'service_revenue': doc_service_rev,
                'doctor_share': doc_total_share,
                'clinic_share': doc_clinic_share,
            })

    # Let's build the Workbook
    wb = openpyxl.Workbook()
    
    # Fonts and Styling
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Segoe UI', size=11, color='000000')
    bold_font = Font(name='Segoe UI', size=11, bold=True, color='000000')
    
    # Colors (Premium Dark theme matches the website)
    header_fill = PatternFill(start_color='1E252B', end_color='1E252B', fill_type='solid') # Slate dark grey
    zebra_fill = PatternFill(start_color='F5F6F8', end_color='F5F6F8', fill_type='solid')
    summary_fill = PatternFill(start_color='E2F3E5', end_color='E2F3E5', fill_type='solid') # Very light green for totals
    title_fill = PatternFill(start_color='16A085', end_color='16A085', fill_type='solid') # Vibrant Teal/Green
    
    # Alignment
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Borders
    thin_side = Side(border_style="thin", color="CCCCCC")
    double_bottom_side = Side(border_style="double", color="333333")
    thin_bottom_side = Side(border_style="thin", color="333333")
    
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_bottom_side)
    total_border = Border(top=thin_bottom_side, bottom=double_bottom_side)

    # Number Format
    currency_format = '#,##0.00 "ج.م"'

    # Helper function to auto-adjust column width
    def adjust_column_widths(ws):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                val_len = len(val)
                # Count double bytes for non-ascii (Arabic)
                for char in val:
                    if ord(char) > 127:
                        val_len += 0.8
                max_len = max(max_len, val_len)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 1. SUMMARY SHEET
    ws_sum = wb.active
    ws_sum.title = "الملخص المالي"
    ws_sum.sheet_view.rightToLeft = True
    
    # Title Block
    ws_sum.merge_cells('A1:D2')
    title_cell = ws_sum['A1']
    title_cell.value = "تقرير الإيرادات والتحليلات المالية للعيادة"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    # Period Info
    ws_sum['A4'] = "الفترة الزمنية:"
    ws_sum['A4'].font = bold_font
    ws_sum['B4'] = f"من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
    ws_sum['B4'].font = normal_font
    
    ws_sum['A5'] = "نوع الكشف:"
    ws_sum['A5'].font = bold_font
    ws_sum['B5'] = "الكل" if exam_type == 'all' else ("نسبة مئوية" if exam_type == 'percentage' else "تقاسم الوقت")
    ws_sum['B5'].font = normal_font

    ws_sum['A6'] = "تاريخ التصدير:"
    ws_sum['A6'].font = bold_font
    ws_sum['B6'] = timezone.now().strftime('%Y-%m-%d %I:%M %p')
    ws_sum['B6'].font = normal_font

    # Financial Summary Table
    ws_sum['A9'] = "البند"
    ws_sum['B9'] = "العدد"
    ws_sum['C9'] = "القيمة / الإيراد"
    ws_sum['A9'].font = header_font
    ws_sum['B9'].font = header_font
    ws_sum['C9'].font = header_font
    ws_sum['A9'].fill = header_fill
    ws_sum['B9'].fill = header_fill
    ws_sum['C9'].fill = header_fill
    ws_sum['A9'].alignment = center_align
    ws_sum['B9'].alignment = center_align
    ws_sum['C9'].alignment = center_align
    
    summary_rows = [
        ("إجمالي الكشوفات (الاستشارات)", appointments.count(), total_app_revenue),
        ("إجمالي الخدمات المقدمة", services.count(), total_service_revenue),
        ("إجمالي الإيرادات الكلية", appointments.count() + services.count(), total_revenue),
        ("إجمالي نصيب الأطباء", "", total_doctor_share),
        ("إجمالي صافي العيادة", "", total_clinic_share),
    ]
    
    current_row = 10
    for title, count, val in summary_rows:
        ws_sum.cell(row=current_row, column=1, value=title).font = bold_font if "إجمالي" in title else normal_font
        ws_sum.cell(row=current_row, column=2, value=count).font = normal_font
        
        val_cell = ws_sum.cell(row=current_row, column=3, value=val)
        val_cell.font = bold_font if "إجمالي" in title else normal_font
        val_cell.number_format = currency_format
        
        # Zebra / Highlight
        for col_idx in range(1, 4):
            c = ws_sum.cell(row=current_row, column=col_idx)
            c.border = cell_border
            if "صافي العيادة" in title:
                c.fill = summary_fill
            elif "الإيرادات الكلية" in title:
                c.fill = zebra_fill
        
        ws_sum.cell(row=current_row, column=1).alignment = right_align
        ws_sum.cell(row=current_row, column=2).alignment = center_align
        ws_sum.cell(row=current_row, column=3).alignment = left_align
        current_row += 1

    # Doctor Performance Table
    current_row += 2
    ws_sum.cell(row=current_row, column=1, value="تحليل أداء الأطباء").font = Font(name='Segoe UI', size=14, bold=True, color='1E252B')
    current_row += 1
    
    doc_headers = ["الطبيب", "نوع التعاقد", "عدد الكشوفات", "إيراد المواعيد", "إيراد الخدمات", "نصيب الطبيب", "نصيب العيادة"]
    for col_idx, text in enumerate(doc_headers, start=1):
        cell = ws_sum.cell(row=current_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    current_row += 1
    start_doc_row = current_row
    for item in doctor_data:
        ws_sum.cell(row=current_row, column=1, value=item['doctor']).font = bold_font
        ws_sum.cell(row=current_row, column=2, value=item['type']).font = normal_font
        ws_sum.cell(row=current_row, column=3, value=item['app_count']).font = normal_font
        ws_sum.cell(row=current_row, column=4, value=item['app_revenue']).number_format = currency_format
        ws_sum.cell(row=current_row, column=5, value=item['service_revenue']).number_format = currency_format
        ws_sum.cell(row=current_row, column=6, value=item['doctor_share']).number_format = currency_format
        ws_sum.cell(row=current_row, column=7, value=item['clinic_share']).number_format = currency_format
        
        # Alignment & Borders
        for col_idx in range(1, 8):
            c = ws_sum.cell(row=current_row, column=col_idx)
            c.border = cell_border
            if col_idx in [1, 2]:
                c.alignment = right_align
            elif col_idx == 3:
                c.alignment = center_align
            else:
                c.alignment = left_align
        current_row += 1
        
    # Doctor Performance Totals
    if doctor_data:
        end_doc_row = current_row - 1
        ws_sum.cell(row=current_row, column=1, value="الإجمالي").font = bold_font
        ws_sum.cell(row=current_row, column=3, value=f"=SUM(C{start_doc_row}:C{end_doc_row})").font = bold_font
        ws_sum.cell(row=current_row, column=4, value=f"=SUM(D{start_doc_row}:D{end_doc_row})").number_format = currency_format
        ws_sum.cell(row=current_row, column=4).font = bold_font
        ws_sum.cell(row=current_row, column=5, value=f"=SUM(E{start_doc_row}:E{end_doc_row})").number_format = currency_format
        ws_sum.cell(row=current_row, column=5).font = bold_font
        ws_sum.cell(row=current_row, column=6, value=f"=SUM(F{start_doc_row}:F{end_doc_row})").number_format = currency_format
        ws_sum.cell(row=current_row, column=6).font = bold_font
        ws_sum.cell(row=current_row, column=7, value=f"=SUM(G{start_doc_row}:G{end_doc_row})").number_format = currency_format
        ws_sum.cell(row=current_row, column=7).font = bold_font
        
        for col_idx in range(1, 8):
            c = ws_sum.cell(row=current_row, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 2]:
                c.alignment = right_align
            elif col_idx == 3:
                c.alignment = center_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_sum)

    # 2. APPOINTMENTS SHEET
    ws_app = wb.create_sheet(title="تفاصيل الكشوفات")
    ws_app.sheet_view.rightToLeft = True
    
    app_headers = ["التاريخ والوقت", "رقم الجلسة", "المريض", "الطبيب", "نوع التعاقد", "التكلفة الكلية", "نصيب الطبيب", "نصيب العيادة", "الحالة"]
    for col_idx, text in enumerate(app_headers, start=1):
        cell = ws_app.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    for row_idx, app in enumerate(appointments, start=2):
        doc = app.doctor
        doc_name = doc.user.get_full_name()
        
        # Share logic
        cost = app.cost
        if doc.examination_type == 'percentage':
            percentage = doc.percentage_value or 0
            doc_share = cost * (percentage / 100)
        elif doc.examination_type == 'time_share':
            price = doc.price_value or 0
            doc_share = price
        else:
            doc_share = cost
            
        clinic_share = cost - doc_share
        
        ws_app.cell(row=row_idx, column=1, value=app.date.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_app.cell(row=row_idx, column=2, value=app.session_number).alignment = center_align
        ws_app.cell(row=row_idx, column=3, value=app.patient.user.get_full_name()).alignment = right_align
        ws_app.cell(row=row_idx, column=4, value=doc_name).alignment = right_align
        ws_app.cell(row=row_idx, column=5, value=doc.get_examination_type_display()).alignment = right_align
        
        cost_cell = ws_app.cell(row=row_idx, column=6, value=cost)
        cost_cell.number_format = currency_format
        cost_cell.alignment = left_align
        
        doc_cell = ws_app.cell(row=row_idx, column=7, value=doc_share)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_app.cell(row=row_idx, column=8, value=clinic_share)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_app.cell(row=row_idx, column=9, value=app.get_status_display()).alignment = center_align
        
        # Styling and zebra striping
        for col_idx in range(1, 10):
            c = ws_app.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
                
    # Appointment Totals
    if appointments.exists():
        total_row_idx = appointments.count() + 2
        ws_app.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_app.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=6).font = bold_font
        ws_app.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=7).font = bold_font
        ws_app.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=8).font = bold_font
        
        for col_idx in range(1, 10):
            c = ws_app.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 2, 9]:
                c.alignment = center_align
            elif col_idx in [3, 4, 5]:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_app)

    # 3. SERVICES SHEET
    ws_ser = wb.create_sheet(title="تفاصيل الخدمات")
    ws_ser.sheet_view.rightToLeft = True
    
    ser_headers = ["تاريخ الخدمة", "المريض", "الخدمة", "الطبيب", "السعر الإجمالي", "نصيب الطبيب", "نصيب العيادة", "بواسطة"]
    for col_idx, text in enumerate(ser_headers, start=1):
        cell = ws_ser.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    for row_idx, ser in enumerate(services, start=2):
        ws_ser.cell(row=row_idx, column=1, value=ser.created_at.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_ser.cell(row=row_idx, column=2, value=ser.patient.user.get_full_name()).alignment = right_align
        ws_ser.cell(row=row_idx, column=3, value=ser.service.name).alignment = right_align
        ws_ser.cell(row=row_idx, column=4, value=ser.doctor.user.get_full_name()).alignment = right_align
        
        price_cell = ws_ser.cell(row=row_idx, column=5, value=ser.service.price)
        price_cell.number_format = currency_format
        price_cell.alignment = left_align
        
        doc_cell = ws_ser.cell(row=row_idx, column=6, value=ser.doctor_amount)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_ser.cell(row=row_idx, column=7, value=ser.clinic_amount)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_ser.cell(row=row_idx, column=8, value=ser.created_by.get_full_name()).alignment = center_align
        
        for col_idx in range(1, 9):
            c = ws_ser.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill

    if services.exists():
        total_row_idx = services.count() + 2
        ws_ser.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_ser.cell(row=total_row_idx, column=5, value=f"=SUM(E2:E{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=5).font = bold_font
        ws_ser.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=6).font = bold_font
        ws_ser.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=7).font = bold_font
        
        for col_idx in range(1, 9):
            c = ws_ser.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 8]:
                c.alignment = center_align
            elif col_idx in [2, 3, 4]:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_ser)

    # 4. ALL APPOINTMENTS SHEET
    ws_all_app = wb.create_sheet(title="كل المواعيد")
    ws_all_app.sheet_view.rightToLeft = True
    
    all_app_headers = ["التاريخ والوقت", "رقم الجلسة", "المريض", "الطبيب", "نوع التعاقد", "التكلفة الكلية", "نصيب الطبيب", "نصيب العيادة", "حالة الموعد"]
    for col_idx, text in enumerate(all_app_headers, start=1):
        cell = ws_all_app.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    # Query ALL appointments in the period (including cancelled)
    all_appointments = Appointment.objects.filter(date__date__range=[start_date, end_date]).order_by('date')
    if exam_type != 'all':
        all_appointments = all_appointments.filter(doctor__examination_type=exam_type)
        
    for row_idx, app in enumerate(all_appointments, start=2):
        doc = app.doctor
        doc_name = doc.user.get_full_name()
        
        cost = app.cost
        if doc.examination_type == 'percentage':
            percentage = doc.percentage_value or 0
            doc_share = cost * (percentage / 100)
        elif doc.examination_type == 'time_share':
            price = doc.price_value or 0
            doc_share = price
        else:
            doc_share = cost
            
        clinic_share = cost - doc_share
        
        ws_all_app.cell(row=row_idx, column=1, value=app.date.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_all_app.cell(row=row_idx, column=2, value=app.session_number).alignment = center_align
        ws_all_app.cell(row=row_idx, column=3, value=app.patient.user.get_full_name()).alignment = right_align
        ws_all_app.cell(row=row_idx, column=4, value=doc_name).alignment = right_align
        ws_all_app.cell(row=row_idx, column=5, value=doc.get_examination_type_display()).alignment = right_align
        
        cost_cell = ws_all_app.cell(row=row_idx, column=6, value=cost)
        cost_cell.number_format = currency_format
        cost_cell.alignment = left_align
        
        doc_cell = ws_all_app.cell(row=row_idx, column=7, value=doc_share)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_all_app.cell(row=row_idx, column=8, value=clinic_share)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_all_app.cell(row=row_idx, column=9, value=app.get_status_display()).alignment = center_align
        
        for col_idx in range(1, 10):
            c = ws_all_app.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
                
    if all_appointments.exists():
        total_row_idx = all_appointments.count() + 2
        ws_all_app.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_all_app.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=6).font = bold_font
        ws_all_app.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=7).font = bold_font
        ws_all_app.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=8).font = bold_font
        
        for col_idx in range(1, 10):
            c = ws_all_app.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 2, 9]:
                c.alignment = center_align
            elif col_idx in [3, 4, 5]:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_all_app)

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response