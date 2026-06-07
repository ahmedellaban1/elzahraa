from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from .models import Doctor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from appointments.models import Appointment
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.utils.crypto import get_random_string
from django.db import transaction
from .forms import DoctorCreationForm, ReceptionistCreationForm, DoctorUpdateForm

@login_required
def get_doctor(request):
    query = request.GET.get('q', '')
    doctor = Doctor.objects.filter(
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(specialization__icontains=query)
    )
    data = []
    for doctor in doctor:
        data.append({
            'id': doctor.id,
            'full_name': doctor.user.get_full_name(),
        })
    return JsonResponse(data, safe=False)
@login_required
@permission_required('appointments.change_appointment', raise_exception=True)
def doctor_dashboard(request):
    # Check if user is a doctor
    if not hasattr(request.user, 'doctor'):
        return render(request, '403.html', status=403)
        
    doctor = request.user.doctor
    
    # Get date from request or default to today
    date_str = request.GET.get('date')
    if date_str:
        try:
            from datetime import datetime
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = timezone.now().date()
    else:
        target_date = timezone.now().date()
    
    # Get appointments for the specific date
    appointments = Appointment.objects.filter(
        doctor=doctor,
        date__date=target_date
    ).order_by('session_number')
    
    context = {
        'page_title': 'لوحة تحكم الطبيب',
        'doctor': doctor,
        'appointments': appointments,
        'target_date': target_date,
    }
    return render(request, 'staff/doctor_dashboard.html', context)


@login_required
@permission_required('staff.add_doctor', raise_exception=True)
def create_doctor(request):

        
    generated_password = None
    
    if request.method == 'POST':
        form = DoctorCreationForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save(commit=False)
                    # Generate a random password
                    generated_password = get_random_string(length=10)
                    user.set_password(generated_password)
                    user.created_by = request.user
                    user.save()
                    # Profile is created in the form save method
                    form.save(commit=True)
                    
                    messages.success(request, f"تم إنشاء حساب الطبيب بنجاح.")
                    # Pass the password to the context for display
                    return render(request, 'staff/create_doctor.html', {
                        'form': form,
                        'generated_password': generated_password,
                        'new_user': user,
                        'page_title': 'تم إنشاء الحساب'
                    })
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء إنشاء الحساب: {str(e)}")
    else:
        form = DoctorCreationForm()
        
    return render(request, 'staff/create_doctor.html', {
        'form': form,
        'page_title': 'إضافة طبيب جديد'
    })


@login_required
@permission_required('staff.change_doctor', raise_exception=True)
def update_doctor(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    user = doctor.user
    if request.method == 'POST':
        form = DoctorUpdateForm(request.POST, instance=user, doctor=doctor)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تحديث بيانات الطبيب {user.get_full_name()} بنجاح.")
            return redirect('staff:doctor_detail', pk=doctor.id)
    else:
        form = DoctorUpdateForm(instance=user, doctor=doctor)
    
    return render(request, 'staff/update_doctor.html', {
        'form': form,
        'doctor': doctor,
        'page_title': f'تعديل بيانات الطبيب: {user.get_full_name()}'
    })


@login_required
@permission_required('staff.view_doctor', raise_exception=True)
def doctor_detail(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    
    # Date Range filtering logic
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    if start_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = today
        
    if end_date_str:
        from datetime import datetime
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = start_date

    # Get appointments and services within range
    appointments = Appointment.objects.filter(
        doctor=doctor,
        date__date__range=[start_date, end_date]
    ).exclude(status='cancelled').order_by('-date')
    
    from services.models import ServiceRecord
    services = ServiceRecord.objects.filter(
        doctor=doctor,
        created_at__date__range=[start_date, end_date]
    ).order_by('-created_at')
    
    # Financial calculations
    from django.db.models import Sum
    doc_app_rev = appointments.aggregate(Sum('cost'))['cost__sum'] or 0
    doc_service_rev = services.aggregate(Sum('service__price'))['service__price__sum'] or 0
    doc_service_share = services.aggregate(Sum('doctor_money'))['doctor_money__sum'] or 0
    
    doc_app_share = appointments.aggregate(Sum('doctor_money'))['doctor_money__sum'] or 0
    doc_app_clinic_share = appointments.aggregate(Sum('clinic_money'))['clinic_money__sum'] or 0
    doc_service_clinic_share = services.aggregate(Sum('clinic_money'))['clinic_money__sum'] or 0
        
    total_revenue = doc_app_rev + doc_service_rev
    doctor_share = doc_app_share + doc_service_share
    clinic_share = doc_app_clinic_share + doc_service_clinic_share

    # Handle temporary password display from session
    new_password = None
    if request.session.get('new_password') and request.session.get('reset_user_id') == doctor.user.id:
        new_password = request.session.pop('new_password')
        request.session.pop('reset_user_id')
    
    context = {
        'page_title': f'ملف الطبيب: {doctor.user.get_full_name()}',
        'doctor': doctor,
        'appointments': appointments,
        'services': services,
        'new_password': new_password,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_revenue': total_revenue,
        'doctor_share': doctor_share,
        'clinic_share': clinic_share,
        'app_count': appointments.count(),
        'service_count': services.count(),
    }
    return render(request, 'staff/doctor_detail.html', context)

@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def reset_doctor_password(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    user = doctor.user
    
    # Generate a new random password
    new_password = get_random_string(length=10)
    user.set_password(new_password)
    user.save()
    
    messages.success(request, f"تم إعادة تعيين كلمة المرور للطبيب {user.get_full_name()} بنجاح.")
    
    # We'll pass the new password in the session to show it once
    request.session['new_password'] = new_password
    request.session['reset_user_id'] = user.id
    
    return redirect('staff:doctor_detail', pk=pk)


@login_required
@permission_required('staff.add_receptionist', raise_exception=True)
def create_receptionist(request):
    generated_password = None
    
    if request.method == 'POST':
        form = ReceptionistCreationForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save(commit=False)
                    # Generate a random password
                    generated_password = get_random_string(length=10)
                    user.set_password(generated_password)
                    user.created_by = request.user
                    user.save()
                    # Profile is created in the form save method
                    form.save(commit=True)
                    
                    messages.success(request, f"تم إنشاء حساب موظف الاستقبال بنجاح.")
                    # Pass the password to the context for display
                    return render(request, 'staff/create_receptionist.html', {
                        'form': form,
                        'generated_password': generated_password,
                        'new_user': user,
                        'page_title': 'تم إنشاء الحساب'
                    })
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء إنشاء الحساب: {str(e)}")
    else:
        form = ReceptionistCreationForm()
        
    return render(request, 'staff/create_receptionist.html', {
        'form': form,
        'page_title': 'إضافة موظف استقبال جديد'
    })


@login_required
@permission_required('staff.view_doctor', raise_exception=True)
def export_doctor_excel(request, pk):
    doctor = get_object_or_404(Doctor, pk=pk)
    
    # Date Range filtering logic
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    if start_date_str:
        from datetime import datetime
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = today
        
    if end_date_str:
        from datetime import datetime
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = start_date

    # Get appointments and services within range
    appointments = Appointment.objects.filter(
        doctor=doctor,
        date__date__range=[start_date, end_date]
    ).exclude(status='cancelled').order_by('date')
    
    from services.models import ServiceRecord
    services = ServiceRecord.objects.filter(
        doctor=doctor,
        created_at__date__range=[start_date, end_date]
    ).order_by('created_at')
    
    # Financial calculations
    from django.db.models import Sum
    d_app_net = appointments.aggregate(Sum('cost'))['cost__sum'] or 0
    app_discount_sum = sum(app.discount.discount_amount for app in appointments if hasattr(app, 'discount'))
    doc_app_rev = d_app_net + app_discount_sum
    
    doc_service_rev = services.aggregate(Sum('service__price'))['service__price__sum'] or 0
    service_discount_sum = sum(ser.discount.discount_amount for ser in services if hasattr(ser, 'discount'))
    
    total_discounts = app_discount_sum + service_discount_sum
    
    doc_service_share = services.aggregate(Sum('doctor_money'))['doctor_money__sum'] or 0
    
    doc_app_share = appointments.aggregate(Sum('doctor_money'))['doctor_money__sum'] or 0
    doc_app_clinic_share = appointments.aggregate(Sum('clinic_money'))['clinic_money__sum'] or 0
    doc_service_clinic_share = services.aggregate(Sum('clinic_money'))['clinic_money__sum'] or 0
        
    total_original_revenue = doc_app_rev + doc_service_rev
    total_net_revenue = total_original_revenue - total_discounts
    doctor_share = doc_app_share + doc_service_share
    clinic_share = doc_app_clinic_share + doc_service_clinic_share

    # Build the Workbook
    wb = openpyxl.Workbook()
    
    # Fonts and Styling
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='Segoe UI', size=11, color='000000')
    bold_font = Font(name='Segoe UI', size=11, bold=True, color='000000')
    
    # Colors
    header_fill = PatternFill(start_color='1E252B', end_color='1E252B', fill_type='solid')
    zebra_fill = PatternFill(start_color='F5F6F8', end_color='F5F6F8', fill_type='solid')
    summary_fill = PatternFill(start_color='E2F3E5', end_color='E2F3E5', fill_type='solid')
    title_fill = PatternFill(start_color='16A085', end_color='16A085', fill_type='solid')
    
    # Alignment
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Borders
    thin_side = Side(border_style="thin", color="CCCCCC")
    double_bottom_side = Side(border_style="double", color="333333")
    thin_bottom_side = Side(border_style="thin", color="333333")
    
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_bottom_side)
    total_border = Border(top=thin_bottom_side, bottom=double_bottom_side)

    currency_format = '#,##0.00 "ج.م"'

    def adjust_column_widths(ws):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                val_len = len(val)
                for char in val:
                    if ord(char) > 127:
                        val_len += 0.8
                max_len = max(max_len, val_len)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 1. SUMMARY SHEET
    ws_sum = wb.active
    ws_sum.title = "ملخص الحسابات"
    ws_sum.sheet_view.rightToLeft = True
    
    ws_sum.merge_cells('A1:D2')
    title_cell = ws_sum['A1']
    title_cell.value = f"التقرير المالي للطبيب: د. {doctor.user.get_full_name()}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    ws_sum['A4'] = "التخصص:"
    ws_sum['A4'].font = bold_font
    ws_sum['B4'] = doctor.specialization or "غير محدد"
    ws_sum['B4'].font = normal_font
    
    ws_sum['A5'] = "نظام التعاقد:"
    ws_sum['A5'].font = bold_font
    ws_sum['B5'] = doctor.get_examination_type_display()
    ws_sum['B5'].font = normal_font

    ws_sum['A6'] = "القيمة المحددة:"
    ws_sum['A6'].font = bold_font
    if doctor.examination_type == 'percentage':
        ws_sum['B6'] = f"{doctor.percentage_value}%"
    elif doctor.examination_type == 'time_share':
        ws_sum['B6'] = f"{doctor.price_value} جنيه للموعد"
    else:
        ws_sum['B6'] = "100%"
    ws_sum['B6'].font = normal_font

    ws_sum['A7'] = "الفترة الزمنية:"
    ws_sum['A7'].font = bold_font
    ws_sum['B7'] = f"من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
    ws_sum['B7'].font = normal_font

    # Financial Summary Table
    ws_sum['A9'] = "البند"
    ws_sum['B9'] = "العدد"
    ws_sum['C9'] = "القيمة / الإيراد"
    ws_sum['A9'].font = header_font
    ws_sum['B9'].font = header_font
    ws_sum['C9'].font = header_font
    ws_sum['A9'].fill = header_fill
    ws_sum['B9'].fill = header_fill
    ws_sum['C9'].fill = header_fill
    ws_sum['A9'].alignment = center_align
    ws_sum['B9'].alignment = center_align
    ws_sum['C9'].alignment = center_align
    
    summary_rows = [
        ("إجمالي الكشوفات", appointments.count(), doc_app_rev),
        ("منها — كشف أول", appointments.filter(type='examination').count(), ""),
        ("منها — متابعة", appointments.filter(type='follow_up').count(), ""),
        ("إجمالي الخدمات المقدمة", services.count(), doc_service_rev),
        ("إجمالي الإيرادات الكلية (قبل الخصم)", appointments.count() + services.count(), total_original_revenue),
        ("إجمالي الخصومات", "", total_discounts),
        ("إجمالي الإيرادات الصافية", "", total_net_revenue),
        ("صافي نصيب الطبيب", "", doctor_share),
        ("صافي نصيب العيادة", "", clinic_share),
    ]
    
    current_row = 10
    for title, count, val in summary_rows:
        ws_sum.cell(row=current_row, column=1, value=title).font = bold_font if "إجمالي" in title or "صافي" in title else normal_font
        ws_sum.cell(row=current_row, column=2, value=count).font = normal_font
        
        val_cell = ws_sum.cell(row=current_row, column=3, value=val)
        val_cell.font = bold_font if "إجمالي" in title or "صافي" in title else normal_font
        val_cell.number_format = currency_format
        
        for col_idx in range(1, 4):
            c = ws_sum.cell(row=current_row, column=col_idx)
            c.border = cell_border
            if "نصيب الطبيب" in title:
                c.fill = summary_fill
            elif "الإيرادات الكلية" in title:
                c.fill = zebra_fill
        
        ws_sum.cell(row=current_row, column=1).alignment = right_align
        ws_sum.cell(row=current_row, column=2).alignment = center_align
        ws_sum.cell(row=current_row, column=3).alignment = left_align
        current_row += 1

    adjust_column_widths(ws_sum)

    # 2. APPOINTMENTS SHEET
    ws_app = wb.create_sheet(title="تفاصيل الكشوفات")
    ws_app.sheet_view.rightToLeft = True
    
    app_headers = ["التاريخ والوقت", "رقم الجلسة", "المريض", "نوع الموعد", "التكلفة الكلية", "الخصم", "نصيب الطبيب", "نصيب العيادة", "الحالة"]
    for col_idx, text in enumerate(app_headers, start=1):
        cell = ws_app.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    for row_idx, app in enumerate(appointments, start=2):
        discount_val = app.discount.discount_amount if hasattr(app, 'discount') else 0
        cost = app.discount.original_amount if hasattr(app, 'discount') else app.cost
        doc_share = app.doctor_money
        clinic_share = app.clinic_money
        
        ws_app.cell(row=row_idx, column=1, value=app.date.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_app.cell(row=row_idx, column=2, value=app.session_number).alignment = center_align
        ws_app.cell(row=row_idx, column=3, value=app.patient.user.get_full_name()).alignment = right_align
        ws_app.cell(row=row_idx, column=4, value='كشف' if app.type == 'examination' else 'متابعة').alignment = center_align
        
        cost_cell = ws_app.cell(row=row_idx, column=5, value=cost)
        cost_cell.number_format = currency_format
        cost_cell.alignment = left_align

        discount_cell = ws_app.cell(row=row_idx, column=6, value=discount_val)
        discount_cell.number_format = currency_format
        discount_cell.alignment = left_align
        
        doc_cell = ws_app.cell(row=row_idx, column=7, value=doc_share)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_app.cell(row=row_idx, column=8, value=clinic_share)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_app.cell(row=row_idx, column=9, value=app.get_status_display()).alignment = center_align
        
        for col_idx in range(1, 10):
            c = ws_app.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
                
    if appointments.exists():
        total_row_idx = appointments.count() + 2
        ws_app.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_app.cell(row=total_row_idx, column=5, value=f"=SUM(E2:E{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=5).font = bold_font
        ws_app.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=6).font = bold_font
        ws_app.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=7).font = bold_font
        ws_app.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{total_row_idx-1})").number_format = currency_format
        ws_app.cell(row=total_row_idx, column=8).font = bold_font
        
        for col_idx in range(1, 10):
            c = ws_app.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 2, 4, 9]:
                c.alignment = center_align
            elif col_idx == 3:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_app)

    # 3. SERVICES SHEET
    ws_ser = wb.create_sheet(title="تفاصيل الخدمات")
    ws_ser.sheet_view.rightToLeft = True
    
    ser_headers = ["التاريخ", "المريض", "الخدمة", "السعر الإجمالي", "الخصم", "نصيب الطبيب", "نصيب العيادة", "بواسطة"]
    for col_idx, text in enumerate(ser_headers, start=1):
        cell = ws_ser.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    for row_idx, ser in enumerate(services, start=2):
        ws_ser.cell(row=row_idx, column=1, value=ser.created_at.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_ser.cell(row=row_idx, column=2, value=ser.patient.user.get_full_name()).alignment = right_align
        ws_ser.cell(row=row_idx, column=3, value=ser.service.name).alignment = right_align
        
        discount_val = ser.discount.discount_amount if hasattr(ser, 'discount') else 0

        price_cell = ws_ser.cell(row=row_idx, column=4, value=ser.service.price)
        price_cell.number_format = currency_format
        price_cell.alignment = left_align

        discount_cell = ws_ser.cell(row=row_idx, column=5, value=discount_val)
        discount_cell.number_format = currency_format
        discount_cell.alignment = left_align
        
        doc_cell = ws_ser.cell(row=row_idx, column=6, value=ser.doctor_money)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_ser.cell(row=row_idx, column=7, value=ser.clinic_money)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_ser.cell(row=row_idx, column=8, value=ser.created_by.get_full_name()).alignment = center_align
        
        for col_idx in range(1, 9):
            c = ws_ser.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill

    if services.exists():
        total_row_idx = services.count() + 2
        ws_ser.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_ser.cell(row=total_row_idx, column=4, value=f"=SUM(D2:D{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=4).font = bold_font
        ws_ser.cell(row=total_row_idx, column=5, value=f"=SUM(E2:E{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=5).font = bold_font
        ws_ser.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=6).font = bold_font
        ws_ser.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_ser.cell(row=total_row_idx, column=7).font = bold_font
        
        for col_idx in range(1, 9):
            c = ws_ser.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 8]:
                c.alignment = center_align
            elif col_idx in [2, 3]:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_ser)

    # 4. ALL APPOINTMENTS SHEET (Doctor specific, including cancelled)
    ws_all_app = wb.create_sheet(title="كل المواعيد")
    ws_all_app.sheet_view.rightToLeft = True
    
    all_app_headers = ["التاريخ والوقت", "رقم الجلسة", "المريض", "نوع الموعد", "التكلفة الكلية", "الخصم", "نصيب الطبيب", "نصيب العيادة", "حالة الموعد"]
    for col_idx, text in enumerate(all_app_headers, start=1):
        cell = ws_all_app.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = header_border
        
    all_appointments = Appointment.objects.filter(
        doctor=doctor,
        date__date__range=[start_date, end_date]
    ).order_by('date')
    
    for row_idx, app in enumerate(all_appointments, start=2):
        discount_val = app.discount.discount_amount if hasattr(app, 'discount') else 0
        cost = app.discount.original_amount if hasattr(app, 'discount') else app.cost
        doc_share = app.doctor_money
        clinic_share = app.clinic_money
        
        ws_all_app.cell(row=row_idx, column=1, value=app.date.strftime('%Y-%m-%d %I:%M %p')).alignment = center_align
        ws_all_app.cell(row=row_idx, column=2, value=app.session_number).alignment = center_align
        ws_all_app.cell(row=row_idx, column=3, value=app.patient.user.get_full_name()).alignment = right_align
        ws_all_app.cell(row=row_idx, column=4, value='كشف' if app.type == 'examination' else 'متابعة').alignment = center_align
        
        cost_cell = ws_all_app.cell(row=row_idx, column=5, value=cost)
        cost_cell.number_format = currency_format
        cost_cell.alignment = left_align

        discount_cell = ws_all_app.cell(row=row_idx, column=6, value=discount_val)
        discount_cell.number_format = currency_format
        discount_cell.alignment = left_align
        
        doc_cell = ws_all_app.cell(row=row_idx, column=7, value=doc_share)
        doc_cell.number_format = currency_format
        doc_cell.alignment = left_align
        
        clinic_cell = ws_all_app.cell(row=row_idx, column=8, value=clinic_share)
        clinic_cell.number_format = currency_format
        clinic_cell.alignment = left_align
        
        ws_all_app.cell(row=row_idx, column=9, value=app.get_status_display()).alignment = center_align
        
        for col_idx in range(1, 10):
            c = ws_all_app.cell(row=row_idx, column=col_idx)
            c.font = normal_font
            c.border = cell_border
            if row_idx % 2 == 1:
                c.fill = zebra_fill
                
    if all_appointments.exists():
        total_row_idx = all_appointments.count() + 2
        ws_all_app.cell(row=total_row_idx, column=1, value="الإجمالي").font = bold_font
        ws_all_app.cell(row=total_row_idx, column=5, value=f"=SUM(E2:E{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=5).font = bold_font
        ws_all_app.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=6).font = bold_font
        ws_all_app.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=7).font = bold_font
        ws_all_app.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{total_row_idx-1})").number_format = currency_format
        ws_all_app.cell(row=total_row_idx, column=8).font = bold_font
        
        for col_idx in range(1, 10):
            c = ws_all_app.cell(row=total_row_idx, column=col_idx)
            c.border = total_border
            c.fill = summary_fill
            if col_idx in [1, 2, 4, 9]:
                c.alignment = center_align
            elif col_idx == 3:
                c.alignment = right_align
            else:
                c.alignment = left_align

    adjust_column_widths(ws_all_app)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    doc_slug = doctor.user.username
    response['Content-Disposition'] = f'attachment; filename="doctor_{doc_slug}_report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response
